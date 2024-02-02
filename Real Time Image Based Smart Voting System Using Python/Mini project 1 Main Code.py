"""
Author      : VETRIVEL.K
Language    : PYTHON
Description : REAL TIME IMAGE BASED SMART VOTING SYSTEM
Module      : IT IS USED TO IDENTIFY ALREADY VOTED PERSON AND STORE SCAN DATA IN WORK SHEET

"""
import tkinter as tk
from tkinter import messagebox
import cv2
from pyzbar.pyzbar import decode
import openpyxl
from openpyxl.styles import Font
import time
# Create the main application window
root = tk.Tk()
root.geometry("450x200")

# Create a new Excel workbook
workbook = openpyxl.Workbook()

# Select the default sheet (Sheet 1)
sheet = workbook.active

# Define your long header text
long_header_text = "VOTER LIST"

# Merge cells for the long header
sheet.merge_cells('A1:C1')  # Merge cells from A1 to C1
header_cell = sheet['A1']  # Get the merged cell
header_cell.value = long_header_text  # Set the header text

# Create a Font object with bold and large font size
font = Font(size=25, bold=True)  # Adjust the size and boldness as needed

# Apply the font settings to the header cell
header_cell.font = font


# Function to start reading barcodes and store in Excel
def bar():
    cap = cv2.VideoCapture(0)
    used_codes = []
    # workbook = openpyxl.Workbook()
    workspace = workbook.active
    while True:
        success, frame = cap.read()
        for code in decode(frame):
            if code.data.decode('utf-8') not in used_codes:
                messagebox.showinfo(" Voter Status ", f"APPROVED TO VOTE")
                print(code.data.decode('utf-8'))
                used_codes.append(code.data.decode('utf-8'))
                time.sleep(2)
                workspace.append([code.data.decode('utf-8')])
            elif code.data.decode('utf-8') in used_codes:
                messagebox.showinfo(" Voter Status ", f"ALREADY VOTED PLEASE CHECK")
                time.sleep(2)
            else:
                pass
        workbook.save(' Voter Data .xlsx')


# create title for project
scan_button = tk.Button(root, text="SMART VOTING SYSTEM", font="arial 25  bold")
scan_button.pack(pady=10)

# Create a button to start scanning and store in Excel
scan_button = tk.Button(root, text="Start Scan", font="arial 25 bold", command=bar)
scan_button.pack()
root.mainloop()

"""-------------------------------------- CODE END -----------------------------------"""
